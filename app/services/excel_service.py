from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict, Any
from app.models.schemas import ReportRow
import io

class ExcelService:
    
    def create_excel_report(self, data: List[Dict[str, Any]]) -> bytes:
        """
        Create Excel report from report data
        """
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Nutresa Report"
        
        # Define headers for the report structure
        headers = [
            "Código de Agente", "Nombre del Agente", "Período de Tiempo", "Variable",
            "Meta Asignada", "Meta Distribuida", "% Meta", "Incentivo Asignado",
            "Incentivo Distribuido", "% Variables Completadas"
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row_idx, report_row in enumerate(data, 2):
            row_data = [
                report_row.get("codigo_agente", ""),
                report_row.get("nombre_agente", ""),
                report_row.get("periodo_tiempo", ""),
                report_row.get("variable", ""),
                report_row.get("meta_asignada", 0),
                report_row.get("meta_distribuida", 0),
                f"{report_row.get('porcentaje_meta', 0)}%" if report_row.get('porcentaje_meta') is not None else "0.00%",
                report_row.get("incentivo_asignado", 0),
                report_row.get("incentivo_distribuido", 0),
                f"{report_row.get('porcentaje_variables_completadas', 0)}%" if report_row.get('porcentaje_variables_completadas') is not None else "0.00%"
            ]
            
            for col, value in enumerate(row_data, 1):
                worksheet.cell(row=row_idx, column=col, value=value)
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        workbook.save(excel_buffer)
        excel_buffer.seek(0)
        
        return excel_buffer.getvalue()
